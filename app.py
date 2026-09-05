import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import re
import io
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Dashboard e Gestão de Riscos", layout="wide")

# Mascaramento de CPF
def mascarar_cpf(val):
    if pd.isna(val):
        return "N/A"
    digits = re.sub(r'\D', '', str(val)).zfill(11)
    if len(digits) >= 11:
        return f"***.***.{digits[-5:-2]}-{digits[-2:]}"
    return "***.***.***-**"

# Regra oficial para a Coluna W (Status no 104)
def obter_status_104(val):
    if pd.isna(val) or str(val).strip() == '' or str(val).strip().upper() == 'NAN':
        return 'A IMPLANTAR'
    v = str(val).strip().upper()
    if 'IMPLANTAD' in v:
        return 'IMPLANTADA'
    return 'A IMPLANTAR'

# Gerador de Relatório Excel Ajustado
def gerar_excel_formatado(df_original):
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DETALHAMENTO POR PARTICIPANTE"
    
    df_exp = df_original.copy()
    
    col_d_name = df_exp.columns[3] if len(df_exp.columns) > 3 else None
    if col_d_name:
        df_exp[col_d_name] = df_exp[col_d_name].fillna('A IMPLANTAR')
        df_exp[col_d_name] = df_exp[col_d_name].apply(lambda x: 'A IMPLANTAR' if pd.isna(x) or str(x).strip() == '' else str(x))

    col_w_name = df_exp.columns[22] if len(df_exp.columns) > 22 else None
    if col_w_name:
        df_exp[col_w_name] = df_exp[col_w_name].apply(obter_status_104)
        
    col_cpf_name = next((c for c in df_exp.columns if 'CPF' in c.upper()), None)
    if col_cpf_name:
        df_exp[col_cpf_name] = df_exp[col_cpf_name].apply(mascarar_cpf)
        
    col_i = df_exp.columns[8] if len(df_exp.columns) > 8 else None
    col_x = df_exp.columns[23] if len(df_exp.columns) > 23 else None
    
    if col_i:
        df_exp[col_i] = pd.to_datetime(df_exp[col_i], errors='coerce').dt.strftime('%d/%m/%Y').fillna('')
    if col_x:
        df_exp[col_x] = pd.to_datetime(df_exp[col_x], errors='coerce').dt.strftime('%d/%m/%Y').fillna('')

    cols_drop = ['DATA_REF', 'ANO', 'MES_ANO', 'TRIMESTRE', 'SEMESTRE', 'STATUS_DASHBOARD']
    df_exp.drop(columns=[c for c in cols_drop if c in df_exp.columns], inplace=True, errors='ignore')
    
    headers = list(df_exp.columns)
    ws.append(headers)
    
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    for row in df_exp.itertuples(index=False):
        ws.append(list(row))
        
    for row_idx in range(2, ws.max_row + 1):
        cell_a = ws.cell(row=row_idx, column=1)
        if cell_a.value is not None and str(cell_a.value).strip() != '':
            try:
                cell_a.value = int(float(cell_a.value))
                cell_a.number_format = '0'
            except: pass

        cell_c = ws.cell(row=row_idx, column=3)
        if cell_c.value is not None and str(cell_c.value).strip() != '':
            try:
                cell_c.value = int(float(cell_c.value))
                cell_c.number_format = '0'
            except: pass

        for col_idx in range(1, ws.max_column + 1):
            if col_idx in [1, 3]: 
                continue
            cell = ws.cell(row=row_idx, column=col_idx)
            if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                cell.number_format = '#,##0.00'
                
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 40)
        
    wb.save(output)
    return output.getvalue()

# Título em Azul Claro estilizado
st.markdown("<h1 style='color: #87CEEB;'>📊 Painel de riscos e pecúlios do Plano FLORIPAPREV</h1>", unsafe_allow_html=True)

# CARREGAMENTO AUTOMÁTICO DO ARQUIVO EXATO NO GITHUB
ARQUIVO_PADRAO = "RELACAO DE MALOTE RECEBIDO SEGURADORA.xlsx"

df = None
if os.path.exists(ARQUIVO_PADRAO):
    try:
        df = pd.read_excel(ARQUIVO_PADRAO)
        st.sidebar.success("✅ Dados carregados automaticamente!")
    except Exception as e:
        st.sidebar.error(f"Erro ao ler o arquivo: {e}")
else:
    st.sidebar.warning(f"Arquivo '{ARQUIVO_PADRAO}' não encontrado no GitHub. Envie abaixo:")
    file_up = st.sidebar.file_uploader("Enviar planilha (.xlsx)", type=["xlsx"])
    if file_up is not None:
        df = pd.read_excel(file_up)

if df is not None:
    df.columns = [str(c).strip() for c in df.columns]
    
    col_nome = next((c for c in df.columns if 'NOME' in c.upper() or 'PROPONENTE' in c.upper()), df.columns[5])
    col_cpf = next((c for c in df.columns if 'CPF' in c.upper()), df.columns[6])
    col_corretor = df.columns[7] # Coluna H (AGENTE MAG)
    col_data = df.columns[8]    # Coluna I (A PARTIR DE)
    col_w_status = df.columns[22] if len(df.columns) > 22 else None # Coluna W
    cols_coberturas = df.columns[9:22] # Colunas J até V
    
    if col_w_status:
        df['STATUS_DASHBOARD'] = df[col_w_status].apply(obter_status_104)
    else:
        df['STATUS_DASHBOARD'] = 'A IMPLANTAR'
    
    # Tratamento de datas e períodos
    df['DATA_REF'] = pd.to_datetime(df[col_data], errors='coerce')
    df['ANO'] = df['DATA_REF'].dt.year.astype(str)
    df['MES_ANO'] = df['DATA_REF'].dt.strftime('%m/%Y')
    df['TRIMESTRE'] = df['DATA_REF'].apply(lambda d: f"{d.year}-Q{(d.month-1)//3 + 1}" if pd.notna(d) else 'N/A')
    df['SEMESTRE'] = df['DATA_REF'].apply(lambda d: f"{d.year} - {1 if d.month <= 6 else 2}º Semestre" if pd.notna(d) else 'N/A')
    
    # Estrutura das 4 Guias Atualizadas
    tab1, tab2, tab3, tab4 = st.tabs([
        "📈 Vendas no Mês, Trimestre e Ano",
        "🏆 Ranking de Produtores (Coluna H)",
        "📋 Relatório dos riscos e pecúlios contratados",
        "💰 Resumo de Contribuições"
    ])
    
    # -------------------------------------------------------------------------
    # GUIA 1: VENDAS GERAIS (Barras Horizontais Coloridas)
    # -------------------------------------------------------------------------
    with tab1:
        st.subheader("Propostas e Riscos Implantados por Período")
        vis_opcao = st.radio("Selecione a escala temporal:", ["Mês", "Trimestre", "Ano"], key="vis_g1", horizontal=True)
        col_p = 'MES_ANO' if vis_opcao == "Mês" else ('TRIMESTRE' if vis_opcao == "Trimestre" else 'ANO')
        
        df_vendas = df.groupby(col_p).size().reset_index(name='Numero de riscos implantados')
        df_vendas = df_vendas[df_vendas[col_p] != 'N/A']
        
        # Ordenação cronológica correta
        if vis_opcao == "Mês":
            df_vendas['TEMP_DATE'] = pd.to_datetime(df_vendas[col_p], format='%m/%Y', errors='coerce')
            df_vendas = df_vendas.sort_values(by='TEMP_DATE', ascending=True).drop(columns=['TEMP_DATE'])
        else:
            df_vendas = df_vendas.sort_values(by=col_p, ascending=True)
        
        # Gráfico de barras horizontais com cores variadas
        fig = px.bar(
            df_vendas, 
            x='Numero de riscos implantados', 
            y=col_p, 
            orientation='h',
            color=col_p,
            color_discrete_sequence=px.colors.qualitative.Vivid,
            text='Numero de riscos implantados',
            title=f"Volume de Riscos Implantados por {vis_opcao}"
        )
        fig.update_layout(showlegend=False)
        st.plotly_chart(fig, use_container_width=True)
        
        # Tabela com as colunas solicitadas
        df_tabela = df_vendas.rename(columns={col_p: 'Mês/Ano'})
        st.dataframe(df_tabela, use_container_width=True)

    # -------------------------------------------------------------------------
    # GUIA 2: RANKING DE PRODUTORES
    # -------------------------------------------------------------------------
    with tab2:
        st.subheader("🏆 Ranking de Produtores (Coluna H)")
        
        vis_ranking = st.radio(
            "Selecione a base temporal para o Ranking:",
            ["Mensal", "Trimestral", "Semestral", "Anual"],
            key="vis_ranking_sel",
            horizontal=True
        )
        
        if vis_ranking == "Mensal":
            col_rank = 'MES_ANO'
            titulo_rank = "Ranking de Vendas por Mês (Do Maior para o Menor)"
        elif vis_ranking == "Trimestral":
            col_rank = 'TRIMESTRE'
            titulo_rank = "Ranking de Vendas por Trimestre (Do Maior para o Menor)"
        elif vis_ranking == "Semestral":
            col_rank = 'SEMESTRE'
            titulo_rank = "Ranking de Vendas por Semestre (Do Maior para o Menor)"
        else:
            col_rank = 'ANO'
            titulo_rank = "Ranking de Vendas por Ano (Do Maior para o Menor)"
            
        st.markdown(f"#### {titulo_rank}")
        
        df_ranking_res = df.groupby([col_rank, col_corretor]).size().reset_index(name='Qtd Vendas')
        df_ranking_res = df_ranking_res[df_ranking_res[col_rank] != 'N/A']
        
        df_ranking_res['TEMP_DATE'] = pd.to_datetime(df_ranking_res[col_rank], format='%m/%Y', errors='coerce')
        df_ranking_res = df_ranking_res.sort_values(by=['TEMP_DATE', 'Qtd Vendas'], ascending=[False, False]).drop(columns=['TEMP_DATE']).reset_index(drop=True)
        
        st.dataframe(df_ranking_res, use_container_width=True)

    # -------------------------------------------------------------------------
    # GUIA 3: RELATÓRIO DOS RISCOS E PECÚLIOS CONTRATADOS
    # -------------------------------------------------------------------------
    with tab3:
        st.subheader("Relatório dos riscos e pecúlios contratados")
        
        lista_participantes = sorted(df[col_nome].dropna().unique().tolist())
        
        if lista_participantes:
            participante_sel = st.selectbox(
                "Busque ou selecione o participante:",
                options=lista_participantes,
                index=0
            )
            
            df_part = df[df[col_nome] == participante_sel].iloc[0]
            status_final = df_part['STATUS_DASHBOARD']
            
            if status_final == 'IMPLANTADA':
                status_html = "<span style='background-color: #D1FAE5; color: #059669; padding: 4px 10px; border-radius: 6px; font-weight: bold;'>IMPLANTADA</span>"
            else:
                status_html = "<span style='background-color: #FEE2E2; color: #DC2626; padding: 4px 10px; border-radius: 6px; font-weight: bold;'>A IMPLANTAR</span>"
            
            st.markdown("---")
            i1, i2, i3 = st.columns(3)
            with i1:
                st.markdown(f"**Proponente:** {df_part[col_nome]}")
                st.markdown(f"**CPF:** `{mascarar_cpf(df_part[col_cpf])}`")
            with i2:
                st.markdown(f"**Proposta:** {df_part.get('PROPOSTA', 'N/A')}")
                st.markdown(f"**Agente/Corretor:** {df_part[col_corretor]}")
            with i3:
                st.markdown(f"**Status:** {status_html}", unsafe_allow_html=True)
                st.markdown(f"**Data da Proposta:** {df_part['MES_ANO']}")
                
            st.markdown("#### Detalhamento de Pecúlios e Riscos (Colunas J a V)")
            
            detalhes_cob = []
            for col_c in cols_coberturas:
                val = df_part[col_c]
                val_num = pd.to_numeric(val, errors='coerce')
                
                if pd.notna(val_num) and val_num != 0:
                    val_fmt = f"R$ {val_num:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                    detalhes_cob.append({"Item / Cobertura": col_c, "Valor Registrado": val_fmt})
                elif pd.notna(val) and str(val).strip() != "" and str(val).upper() != "NÃO":
                    detalhes_cob.append({"Item / Cobertura": col_c, "Valor Registrado": str(val)})
                    
            if detalhes_cob:
                st.table(pd.DataFrame(detalhes_cob))
            else:
                st.info("Nenhuma cobertura ativa para o participante selecionado.")
                
            st.markdown("---")
            excel_bytes = gerar_excel_formatado(df)
            st.download_button(
                label="📥 Exportar Relatório Completo Ajustado em Excel",
                data=excel_bytes,
                file_name="Relatorio_Riscos_Peculios_Completo.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.warning("Nenhum participante encontrado na base de dados.")

    # -------------------------------------------------------------------------
    # GUIA 4: RESUMO E FILTRO DE CONTRIBUICOES
    # -------------------------------------------------------------------------
    with tab4:
        st.subheader("💰 Somatório e Filtro por Tipo de Contribuição")
        
        c_j = df.columns[9] if len(df.columns) > 9 else None   # 2554 - PECÚLIO POR MORTE PÚBLICO PARTICIPANTE
        c_l = df.columns[11] if len(df.columns) > 11 else None # 2554 - PECÚLIO POR MORTE PÚBLICO PATROCINADORA
        c_n = df.columns[13] if len(df.columns) > 13 else None # 2553 - PECÚLIO POR INVALIDEZ PÚBLICO PARTICIPANTE
        c_p = df.columns[15] if len(df.columns) > 15 else None # 2553 - PECÚLIO POR INVALIDEZ PÚBLICO PATROCINADORA
        c_r = df.columns[17] if len(df.columns) > 17 else None # 2029 - ADICIONAL PECÚLIO POR MORTE PÚBLICO
        c_t = df.columns[19] if len(df.columns) > 19 else None # 2030 - ADICIONAL PECÚLIO POR INVALIDEZ PÚBLICO
        
        dict_contribs = {
            "2554 - Pecúlio Morte (Participante)": c_j,
            "2554 - Pecúlio Morte (Patrocinadora)": c_l,
            "2553 - Pecúlio Invalidez (Participante)": c_n,
            "2553 - Pecúlio Invalidez (Patrocinadora)": c_p,
            "2029 - Adicional Pecúlio Morte": c_r,
            "2030 - Adicional Pecúlio Invalidez": c_t
        }
        
        somas = {}
        for nome_amigavel, col_nome_real in dict_contribs.items():
            if col_nome_real and col_nome_real in df.columns:
                df[col_nome_real] = pd.to_numeric(df[col_nome_real], errors='coerce').fillna(0)
                somas[nome_amigavel] = df[col_nome_real].sum()
            else:
                somas[nome_amigavel] = 0.0

        total_geral = sum(somas.values())

        st.markdown("#### 📊 Quadro Geral de Somatórios por Contribuição")
        
        m1, m2, m3 = st.columns(3)
        with m1:
            st.metric("2554 - Morte (Participante)", f"R$ {somas['2554 - Pecúlio Morte (Participante)']:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
            st.metric("2553 - Invalidez (Participante)", f"R$ {somas['2553 - Pecúlio Invalidez (Participante)']:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
        with m2:
            st.metric("2554 - Morte (Patrocinadora)", f"R$ {somas['2554 - Pecúlio Morte (Patrocinadora)']:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
            st.metric("2553 - Invalidez (Patrocinadora)", f"R$ {somas['2553 - Pecúlio Invalidez (Patrocinadora)']:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
        with m3:
            st.metric("2029 - Adicional Morte", f"R$ {somas['2029 - Adicional Pecúlio Morte']:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
            st.metric("2030 - Adicional Invalidez", f"R$ {somas['2030 - Adicional Pecúlio Invalidez']:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))

        st.markdown("---")
        st.markdown(f"### 🏆 **TOTALIZADOR GERAL:** R$ {total_geral:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
        st.markdown("---")

        st.subheader("🔍 Filtrar Participantes por Tipo de Contribuição")
        tipo_filtro_sel = st.selectbox("Selecione o tipo de contribuição para filtrar:", options=list(dict_contribs.keys()))
        
        col_selecionada = dict_contribs[tipo_filtro_sel]
        if col_selecionada and col_selecionada in df.columns:
            df_filtrado_contrib = df[df[col_selecionada] > 0].copy()
            
            st.markdown(f"Exibindo **{len(df_filtrado_contrib)}** registros com valores em: `{tipo_filtro_sel}`")
            
            if not df_filtrado_contrib.empty:
                cols_mostrar = [col_nome, col_cpf, col_corretor, col_selecionada]
                df_exibicao = df_filtrado_contrib[cols_mostrar].copy()
                df_exibicao[col_cpf] = df_exibicao[col_cpf].apply(mascarar_cpf)
                df_exibicao[col_selecionada] = df_exibicao[col_selecionada].apply(lambda v: f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
                
                st.dataframe(df_exibicao, use_container_width=True)
            else:
                st.info("Nenhum registro encontrado com valor maior que zero para esta contribuição.")
else:
    st.error("O arquivo 'RELACAO DE MALOTE RECEBIDO SEGURADORA.xlsx' não foi encontrado no repositório do GitHub. Certifique-se de enviá-lo para a mesma pasta do app.py.")
